import openpyxl

class Excel_Grabber:
    def __init__(self):
        self.data = openpyxl.load_workbook('Residential Prog Code Ref.xlsx')['Sheet1']
        self.Max_TOSL = [self.data.cell(row=i, column=6).value for i in range(4, 17, 2)]
        self.Max_PSO = [self.data.cell(row=i, column=3).value for i in range(4, 17, 2)]
        self.Max_ISA = [self.data.cell(row=i, column=4).value for i in range(4, 17, 2)]
        self.Min_USA = [self.data.cell(row=i, column=5).value for i in range(4, 17, 2)]
        self.FMax_TOSL = [self.data.cell(row=i, column=6).value for i in range(5, 16, 2)]
        self.FMax_PSO = [self.data.cell(row=i, column=3).value for i in range(5, 16, 2)]
        self.FMax_ISA = [self.data.cell(row=i, column=4).value for i in range(5, 16, 2)]
        self.FMin_USA = [self.data.cell(row=i, column=5).value for i in range(5, 16, 2)]
        self.Setback_R1 = [self.data.cell(row=i, column=9).value for i in range(5, 8)]
        self.Setback_basic_R2 = [self.data.cell(row=i, column=10).value for i in range(5, 8)]
        self.Setback_max_R2 = [self.data.cell(row=i, column=11).value for i in range(5, 8)]
        self.Setback_basic_R3 = [self.data.cell(row=i, column=12).value for i in range(5, 8)]
        self.Setback_max_R3 = [self.data.cell(row=i, column=13).value for i in range(5, 8)]
        self.Setback_R4 = [self.data.cell(row=i, column=14).value for i in range(5, 8)]
        self.Setback_R5 = [self.data.cell(row=i, column=15).value for i in range(5, 8)]
        self.Open_Space_Pct = [self.data.cell(row=i, column=3).value for i in range(20, 27)]